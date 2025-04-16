from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import requests
import pandas as pd
import pytesseract
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import cv2
from selenium.common.exceptions import StaleElementReferenceException
from tqdm import tqdm  # Added for progress tracking

pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'  # Common path in Linux environments


# Globals for Excel
wb = Workbook()
ws = wb.active
ws.title = "Result Data"

# Excel header styling
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_font = Font(bold=True)

def check_captcha():
    wait = WebDriverWait(driver, 1)
    time.sleep(0.1)
    captcha_input = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_TextBox1"]')))
    captcha_input.clear()
    lst = [img.get_attribute('src') for img in driver.find_elements(By.TAG_NAME, 'img')]
    src = lst[1]
    response = requests.get(src)
    if response.status_code == 200:
        with open("sample.jpg", 'wb') as f:
            f.write(response.content)

    img = cv2.imread('sample.jpg', cv2.IMREAD_ANYCOLOR)
    text = pytesseract.image_to_string(img).replace(" ", "").upper()
    captcha_input.send_keys(text)
    time.sleep(0.1)
    try:
        view_result_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnviewresult"]')))
        view_result_btn.click()
    except:
        time.sleep(0.1)
    try:
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(0.1)
        check_captcha()
    except:
        get_result()

sno = 1
flag = 1

def head():
    wait = WebDriverWait(driver, 1)
    time.sleep(0.1)
    subject = ["S.No.", "SGPA", "CGPA", "RESULT"]
    global flag
    flag = 0
    ws.append(subject)
    for col_num in range(1, len(subject) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

def get_result():
    global sno
    wait = WebDriverWait(driver, 2)
    det = [sno]
    try:
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(0.1)
        reset_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnReset"]')))
        reset_btn.click()
        check_captcha()
    except:
        pass
    try:
        time.sleep(0.1)
        sgpa = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_lblSGPA"]')))
        cgpa = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_lblcgpa"]')))
        result = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_lblResultNewGrading"]')))

        det.extend([sgpa.text, cgpa.text, result.text])
        ws.append(det)
        sno += 1

        reset_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnReset"]')))  # Fixed typo in XPath
        reset_btn.click()
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtrollno"]')))
        print(f"Processed: CGPA={det[2]}, RESULT={det[3]}")  # Updated print for clarity
    except:
        try:
            reset_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnReset"]')))
            reset_btn.click()
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtrollno"]')))
        except:
            check_captcha()

def process_roll_numbers(file_path, course, semester, output_folder):
    global flag_csv, str_filename, csv_directory, tab1, var, num_sem

    class DummyTab:
        filename = file_path
        directory = output_folder

    class DummyVar:
        def __init__(self, value):
            self.value = value
        def get(self):
            return self.value
        def set(self, v):
            self.value = v

    tab1 = DummyTab()
    var = DummyVar(int(course))
    num_sem = DummyVar(int(semester))
    flag_csv = 1
    str_filename = "results"
    csv_directory = output_folder

    # Validate CSV format
    try:
        data = pd.read_csv(file_path)
        if data.empty or len(data.columns) < 1:
            raise ValueError("CSV is empty or invalid")
        # Rename first column to ensure consistency
        data.columns = ['Roll Number'] + list(data.columns[1:])
        data.to_csv(file_path, index=False)
    except Exception as e:
        raise ValueError(f"Invalid CSV file: {str(e)}")

    sel()
    return os.path.join(csv_directory, str_filename + ".xlsx")

def sel():
    if flag_csv == 1:
        global driver
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1280,800')
        chrome_options.add_argument('--remote-debugging-port=9222')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-software-rasterizer')

        driver = webdriver.Chrome(options=chrome_options)
        driver.get('http://result.rgpv.ac.in/result/programselect.aspx?id=$%')
        wait = WebDriverWait(driver, 2)

        course_btn = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="radlstProgram_{var.get()}"]')))
        course_btn.click()

        with open(str(tab1.filename), 'r') as f:
            data = pd.read_csv(f)
            roll = data.values.tolist()

        head()

        # Wrap roll number loop with tqdm for progress tracking
        for num in tqdm(range(len(roll)), desc="Processing Roll Numbers", unit="student"):
            max_retries = 2
            for attempt in range(max_retries):
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtrollno"]')))
                    roll_input = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtrollno"]')))
                    roll_input.clear()
                    roll_input.send_keys(str(roll[num][0]))  # Convert to string to handle any format
                    print(f"Processing roll number: {roll[num][0]}")  # Status per roll number
                    break
                except StaleElementReferenceException:
                    if attempt < max_retries - 1:
                        time.sleep(0.1)
                        continue
                    else:
                        print(f"Failed for {roll[num][0]} after {max_retries} attempts")
                        raise

            sem_option = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="ctl00_ContentPlaceHolder1_drpSemester"]/option[{num_sem.get()}]')))
            sem_option.click()
            check_captcha()

        driver.quit()
        save_path = os.path.join(csv_directory, str_filename + ".xlsx")
        wb.save(save_path)
        print(f"Saved to {save_path}")
    else:
        print("Invalid Data Entry")